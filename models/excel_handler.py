import openpyxl
import os
from datetime import datetime

class ExcelHandler:
    def __init__(self, base_dir='static/excel_data'):
        # Crear la carpeta si no existe
        os.makedirs(base_dir, exist_ok=True)
        
        # Crear un nombre de archivo con la fecha y hora actuales
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.file_path = os.path.join(base_dir, f'data_user_{timestamp}.xlsx')

    def create_or_load_workbook(self):
        if os.path.exists(self.file_path):
            # Load existing workbook
            workbook = openpyxl.load_workbook(self.file_path)
        else:
            # Create a new workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Add headers if the file is new
            sheet.append(['Name', 'Age', 'Email', 'Phone', 'Address'])
            workbook.save(self.file_path)
        return workbook

    def save_data(self, data):
        workbook = self.create_or_load_workbook()
        sheet = workbook.active
        # Append the new data
        sheet.append(data)
        workbook.save(self.file_path)
        
    
